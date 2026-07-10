from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import sqlite3
import tempfile
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

from .automation_models import AutomationColumn, AutomationTask

APP_DATA_DIR = Path.home() / ".qq_message_manager"
AUTOMATION_ROOT = APP_DATA_DIR / "automation_workspace"
AUTOMATION_STATE_DB = APP_DATA_DIR / "automation_state.sqlite3"
MAX_EXISTING_RECORDS_FOR_AI = 300


class AutomationStorageError(RuntimeError):
    pass


class AutomationStateStore:
    def __init__(self, path: Path = AUTOMATION_STATE_DB) -> None:
        self.path = path
        self._initialize()

    def _connect(self) -> sqlite3.Connection:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        connection = sqlite3.connect(self.path, timeout=15)
        connection.row_factory = sqlite3.Row
        return connection

    def _initialize(self) -> None:
        with self._connect() as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS task_state (
                    task_id TEXT PRIMARY KEY,
                    checkpoint_time TEXT NOT NULL DEFAULT '',
                    checkpoint_message_id TEXT NOT NULL DEFAULT '',
                    last_run_started_at TEXT NOT NULL DEFAULT '',
                    last_run_finished_at TEXT NOT NULL DEFAULT '',
                    last_status TEXT NOT NULL DEFAULT '',
                    last_error TEXT NOT NULL DEFAULT '',
                    retry_count INTEGER NOT NULL DEFAULT 0
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS processed_messages (
                    task_id TEXT NOT NULL,
                    message_key TEXT NOT NULL,
                    processed_at TEXT NOT NULL,
                    PRIMARY KEY (task_id, message_key)
                )
                """
            )

    def state(self, task_id: str) -> dict[str, Any]:
        with self._connect() as connection:
            row = connection.execute("SELECT * FROM task_state WHERE task_id = ?", (task_id,)).fetchone()
        return dict(row) if row is not None else {}

    def checkpoint_time(self, task: AutomationTask) -> datetime:
        raw = self.state(task.task_id).get("checkpoint_time") or task.created_at
        try:
            return datetime.fromisoformat(str(raw)).replace(tzinfo=None)
        except ValueError:
            return task.created_datetime

    def processed_keys(self, task_id: str, keys: list[str]) -> set[str]:
        cleaned = [key for key in dict.fromkeys(keys) if key]
        if not cleaned:
            return set()
        result: set[str] = set()
        with self._connect() as connection:
            for start in range(0, len(cleaned), 500):
                chunk = cleaned[start : start + 500]
                placeholders = ",".join("?" for _ in chunk)
                rows = connection.execute(
                    f"SELECT message_key FROM processed_messages WHERE task_id = ? AND message_key IN ({placeholders})",
                    (task_id, *chunk),
                ).fetchall()
                result.update(str(row[0]) for row in rows)
        return result

    def mark_started(self, task_id: str) -> None:
        now = datetime.now().isoformat()
        with self._connect() as connection:
            connection.execute(
                """
                INSERT INTO task_state(task_id, last_run_started_at, last_status, last_error)
                VALUES(?, ?, 'running', '')
                ON CONFLICT(task_id) DO UPDATE SET
                    last_run_started_at = excluded.last_run_started_at,
                    last_status = 'running',
                    last_error = ''
                """,
                (task_id, now),
            )

    def mark_success(
        self,
        task_id: str,
        checkpoint_time: datetime,
        checkpoint_message_id: str,
        message_keys: list[str],
        status: str = "success",
    ) -> None:
        now = datetime.now().isoformat()
        with self._connect() as connection:
            connection.execute(
                """
                INSERT INTO task_state(
                    task_id, checkpoint_time, checkpoint_message_id,
                    last_run_finished_at, last_status, last_error, retry_count
                ) VALUES(?, ?, ?, ?, ?, '', 0)
                ON CONFLICT(task_id) DO UPDATE SET
                    checkpoint_time = excluded.checkpoint_time,
                    checkpoint_message_id = excluded.checkpoint_message_id,
                    last_run_finished_at = excluded.last_run_finished_at,
                    last_status = excluded.last_status,
                    last_error = '',
                    retry_count = 0
                """,
                (task_id, checkpoint_time.isoformat(), checkpoint_message_id, now, status),
            )
            connection.executemany(
                "INSERT OR IGNORE INTO processed_messages(task_id, message_key, processed_at) VALUES(?, ?, ?)",
                [(task_id, key, now) for key in dict.fromkeys(message_keys) if key],
            )

    def mark_failure(self, task_id: str, error: str, retry_count: int) -> None:
        now = datetime.now().isoformat()
        with self._connect() as connection:
            connection.execute(
                """
                INSERT INTO task_state(task_id, last_run_finished_at, last_status, last_error, retry_count)
                VALUES(?, ?, 'failed', ?, ?)
                ON CONFLICT(task_id) DO UPDATE SET
                    last_run_finished_at = excluded.last_run_finished_at,
                    last_status = 'failed',
                    last_error = excluded.last_error,
                    retry_count = excluded.retry_count
                """,
                (task_id, now, error[:2000], retry_count),
            )

    def delete_task(self, task_id: str) -> None:
        with self._connect() as connection:
            connection.execute("DELETE FROM task_state WHERE task_id = ?", (task_id,))
            connection.execute("DELETE FROM processed_messages WHERE task_id = ?", (task_id,))


def task_workspace(task_id: str) -> Path:
    safe_id = re.sub(r"[^A-Za-z0-9_.-]", "_", task_id)[:100]
    root = (AUTOMATION_ROOT / safe_id).resolve()
    root.mkdir(parents=True, exist_ok=True)
    if AUTOMATION_ROOT.resolve() not in root.parents:
        raise AutomationStorageError("任务工作区路径越界")
    return root


def artifact_path(task: AutomationTask, work_date: date) -> Path:
    task.normalize()
    rendered = task.file_name_template.format(
        date=work_date.isoformat(),
        task_name=_safe_filename(task.name),
        task_id=task.task_id,
    )
    rendered = _safe_filename(rendered)
    expected_suffix = f".{task.file_format}"
    if not rendered.lower().endswith(expected_suffix):
        rendered = Path(rendered).stem + expected_suffix
    path = (task_workspace(task.task_id) / rendered).resolve()
    if task_workspace(task.task_id) not in path.parents:
        raise AutomationStorageError("文件路径越界")
    return path


def records_path(path: Path) -> Path:
    return path.with_name(path.name + ".records.json")


def load_records(path: Path) -> list[dict[str, Any]]:
    sidecar = records_path(path)
    try:
        raw = json.loads(sidecar.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(raw, list):
        return []
    records: list[dict[str, Any]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        record_id = str(item.get("record_id") or "")
        values = item.get("values")
        if not record_id or not isinstance(values, dict):
            continue
        records.append(
            {
                "record_id": record_id,
                "values": {str(key): value for key, value in values.items()},
                "source_message_ids": [str(value) for value in item.get("source_message_ids", []) if str(value)],
                "created_at": str(item.get("created_at") or ""),
                "updated_at": str(item.get("updated_at") or ""),
                "dedup_key": str(item.get("dedup_key") or ""),
            }
        )
    return records


def save_records(path: Path, records: list[dict[str, Any]]) -> None:
    payload = json.dumps(records, ensure_ascii=False, indent=2, default=str)
    _atomic_write_text(records_path(path), payload)


def records_for_ai(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "record_id": record.get("record_id"),
            "values": record.get("values", {}),
            "source_message_ids": record.get("source_message_ids", []),
        }
        for record in records[-MAX_EXISTING_RECORDS_FOR_AI:]
    ]


def apply_operations(
    task: AutomationTask,
    records: list[dict[str, Any]],
    operations: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    task.normalize()
    columns = {column.name: column for column in task.columns}
    by_id = {str(record.get("record_id")): record for record in records}
    dedup_index = {
        str(record.get("dedup_key")): record
        for record in records
        if str(record.get("dedup_key") or "")
    }
    stats = {"inserted": 0, "updated": 0, "ignored": 0}
    now = datetime.now().isoformat()

    for operation in operations[:500]:
        if not isinstance(operation, dict):
            stats["ignored"] += 1
            continue
        action = str(operation.get("action") or "").strip().lower()
        values_raw = operation.get("values")
        if not isinstance(values_raw, dict):
            values_raw = {}
        source_ids = [str(value) for value in operation.get("source_message_ids", []) if str(value)]

        if action == "insert":
            values = _validated_values(columns, values_raw, for_insert=True)
            if values is None:
                stats["ignored"] += 1
                continue
            dedup_key = _dedup_key(task, values)
            existing = dedup_index.get(dedup_key) if dedup_key else None
            if existing is not None:
                changed = _update_record(existing, columns, values, source_ids, now)
                stats["updated" if changed else "ignored"] += 1
                continue
            record = {
                "record_id": f"row_{uuid.uuid4().hex[:16]}",
                "values": values,
                "source_message_ids": list(dict.fromkeys(source_ids)),
                "created_at": now,
                "updated_at": now,
                "dedup_key": dedup_key,
            }
            records.append(record)
            by_id[record["record_id"]] = record
            if dedup_key:
                dedup_index[dedup_key] = record
            stats["inserted"] += 1
            continue

        if action == "update":
            record_id = str(operation.get("record_id") or "").strip()
            record = by_id.get(record_id)
            if record is None:
                stats["ignored"] += 1
                continue
            values = _validated_values(columns, values_raw, for_insert=False)
            if values is None:
                stats["ignored"] += 1
                continue
            changed = _update_record(record, columns, values, source_ids, now)
            new_key = _dedup_key(task, record.get("values", {}))
            record["dedup_key"] = new_key
            if new_key:
                dedup_index[new_key] = record
            stats["updated" if changed else "ignored"] += 1
            continue

        stats["ignored"] += 1

    return records, stats


def write_artifact(task: AutomationTask, work_date: date, records: list[dict[str, Any]]) -> Path:
    path = artifact_path(task, work_date)
    path.parent.mkdir(parents=True, exist_ok=True)
    if task.file_format == "xlsx":
        _write_xlsx(path, task, records)
    elif task.file_format == "csv":
        _write_csv(path, task, records)
    elif task.file_format == "json":
        _write_json(path, task, records)
    elif task.file_format == "md":
        _write_markdown(path, task, records)
    else:
        raise AutomationStorageError(f"不支持的文件格式：{task.file_format}")
    save_records(path, records)
    return path


def ensure_empty_artifact(task: AutomationTask, work_date: date) -> Path:
    path = artifact_path(task, work_date)
    if path.exists():
        return path
    return write_artifact(task, work_date, [])


def delete_artifact_bundle(path: Path) -> None:
    for target in (path, records_path(path)):
        try:
            target.unlink(missing_ok=True)
        except OSError as exc:
            raise AutomationStorageError(f"删除归档文件失败：{exc}") from exc


def message_key(message: Any) -> str:
    message_id = str(getattr(message, "message_id", "") or "").strip()
    if message_id:
        return f"id:{message_id}"
    raw = "|".join(
        [
            str(getattr(message, "session_id", "")),
            str(getattr(message, "sender_id", "")),
            str(int(getattr(message, "timestamp").timestamp())),
            str(getattr(message, "text", "")),
        ]
    )
    return "sha:" + hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def _validated_values(
    columns: dict[str, AutomationColumn],
    raw: dict[str, Any],
    *,
    for_insert: bool,
) -> dict[str, Any] | None:
    values: dict[str, Any] = {}
    for name, column in columns.items():
        if name in raw:
            if not for_insert and not column.ai_update:
                continue
            values[name] = _coerce_value(column, raw[name])
        elif for_insert:
            values[name] = _coerce_value(column, column.default) if column.default != "" else ""
        if for_insert and column.required and _is_empty(values.get(name)):
            return None
    return values


def _coerce_value(column: AutomationColumn, value: Any) -> Any:
    if column.value_type == "number":
        try:
            return float(value) if "." in str(value) else int(value)
        except (TypeError, ValueError):
            return ""
    if column.value_type == "boolean":
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "是", "启用"}
    if column.value_type == "datetime":
        text = str(value or "").strip()
        try:
            return datetime.fromisoformat(text).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            return text[:40]
    text = str(value or "").strip()
    if column.value_type == "enum" and column.enum_values:
        return text if text in column.enum_values else (column.default if column.default in column.enum_values else column.enum_values[0])
    return text[:4000]


def _update_record(
    record: dict[str, Any],
    columns: dict[str, AutomationColumn],
    values: dict[str, Any],
    source_ids: list[str],
    now: str,
) -> bool:
    current = record.setdefault("values", {})
    changed = False
    for name, value in values.items():
        column = columns.get(name)
        if column is None or not column.ai_update:
            continue
        if current.get(name) != value:
            current[name] = value
            changed = True
    merged_sources = list(dict.fromkeys([*record.get("source_message_ids", []), *source_ids]))
    if merged_sources != record.get("source_message_ids", []):
        record["source_message_ids"] = merged_sources
        changed = True
    if changed:
        record["updated_at"] = now
    return changed


def _dedup_key(task: AutomationTask, values: dict[str, Any]) -> str:
    if not task.dedup_fields:
        return ""
    parts = [str(values.get(name, "")).strip().lower() for name in task.dedup_fields]
    if not any(parts):
        return ""
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def _write_xlsx(path: Path, task: AutomationTask, records: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise AutomationStorageError("缺少 openpyxl，请先执行 pip install -r requirements.txt") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = task.sheet_name
    headers = [column.name for column in task.columns]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    for record in records:
        values = record.get("values", {})
        sheet.append([values.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    for index, header in enumerate(headers, start=1):
        width = max(12, min(50, len(header) * 2 + 4))
        sheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = width

    metadata = workbook.create_sheet("_QQMM_META")
    metadata.sheet_state = "hidden"
    metadata.append(["record_id", "source_message_ids", "created_at", "updated_at", "dedup_key"])
    for record in records:
        metadata.append(
            [
                record.get("record_id", ""),
                json.dumps(record.get("source_message_ids", []), ensure_ascii=False),
                record.get("created_at", ""),
                record.get("updated_at", ""),
                record.get("dedup_key", ""),
            ]
        )
    temp = path.with_name(path.name + f".{uuid.uuid4().hex}.tmp")
    workbook.save(temp)
    os.replace(temp, path)


def _write_csv(path: Path, task: AutomationTask, records: list[dict[str, Any]]) -> None:
    headers = [column.name for column in task.columns]
    temp = path.with_name(path.name + f".{uuid.uuid4().hex}.tmp")
    with temp.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for record in records:
            values = record.get("values", {})
            writer.writerow({header: values.get(header, "") for header in headers})
    os.replace(temp, path)


def _write_json(path: Path, task: AutomationTask, records: list[dict[str, Any]]) -> None:
    del task
    visible = [record.get("values", {}) for record in records]
    _atomic_write_text(path, json.dumps(visible, ensure_ascii=False, indent=2, default=str))


def _write_markdown(path: Path, task: AutomationTask, records: list[dict[str, Any]]) -> None:
    headers = [column.name for column in task.columns]
    lines = [
        "| " + " | ".join(_md_escape(header) for header in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for record in records:
        values = record.get("values", {})
        lines.append("| " + " | ".join(_md_escape(values.get(header, "")) for header in headers) + " |")
    _atomic_write_text(path, "\n".join(lines) + "\n")


def _atomic_write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=path.name + ".", suffix=".tmp", dir=str(path.parent))
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_name, path)
    finally:
        try:
            Path(temp_name).unlink(missing_ok=True)
        except OSError:
            pass


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[<>:\"/\\|?*\x00-\x1f]", "_", str(value or "")).strip(" .")
    return cleaned[:180] or "automation_output"


def _md_escape(value: Any) -> str:
    return str(value if value is not None else "").replace("|", "\\|").replace("\n", "<br>")


def _is_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""
