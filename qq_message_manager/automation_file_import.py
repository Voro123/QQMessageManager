from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

SUPPORTED_SUFFIXES = {".xlsx", ".csv", ".json", ".md"}
META_SHEET_NAME = "_QQMM_META"
MAX_IMPORTED_ROWS = 10000


class AutomationFileImportError(RuntimeError):
    pass


def install_automation_file_import(automation_module: Any, storage_module: Any) -> None:
    """让定时任务真正读取用户已有或手动修改过的工作区文件。"""
    if getattr(automation_module, "_stage2_file_import_installed", False):
        return

    task_by_workspace: dict[str, Any] = {}
    original_artifact_path = automation_module.artifact_path
    original_load_records = automation_module.load_records

    def artifact_path_with_registry(task: Any, work_date: Any) -> Path:
        path = Path(original_artifact_path(task, work_date))
        task_by_workspace[str(path.resolve().parent)] = task
        return path

    def load_records_with_file_refresh(path: Path) -> list[dict[str, Any]]:
        artifact = Path(path)
        task = task_by_workspace.get(str(artifact.resolve().parent))
        stored = original_load_records(artifact)
        if not artifact.is_file():
            return stored
        if _sidecar_is_current(storage_module, artifact):
            return stored

        records, recognized = read_artifact_records(artifact, task, storage_module)
        if not recognized:
            return stored
        storage_module.save_records(artifact, records)
        return records

    automation_module.artifact_path = artifact_path_with_registry
    automation_module.load_records = load_records_with_file_refresh
    automation_module._automation_task_by_workspace = task_by_workspace
    automation_module._stage2_file_import_installed = True


def import_user_file(
    task: Any,
    source: Path,
    target: Path,
    automation_module: Any,
    storage_module: Any,
) -> tuple[Path, int]:
    """由用户显式选择文件后，把副本安全导入当前任务工作区。"""
    source = Path(source).expanduser().resolve()
    target = Path(target).expanduser().resolve()
    suffix = source.suffix.lower()
    expected = f".{str(task.file_format).lower()}"
    if suffix not in SUPPORTED_SUFFIXES:
        raise AutomationFileImportError(f"不支持的文件格式：{suffix or '无扩展名'}")
    if suffix != expected:
        raise AutomationFileImportError(
            f"所选文件是 {suffix}，当前任务配置为 {expected}；请先修改任务文件格式"
        )
    if not source.is_file():
        raise AutomationFileImportError("所选文件不存在")

    workspace = storage_module.task_workspace(task.task_id).resolve()
    if workspace not in target.parents:
        raise AutomationFileImportError("导入目标不在当前任务工作区")
    workspace.mkdir(parents=True, exist_ok=True)

    fd, temp_name = tempfile.mkstemp(
        prefix=target.name + ".",
        suffix=".importing",
        dir=str(workspace),
    )
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        shutil.copy2(source, temp_path)
        os.replace(temp_path, target)
    finally:
        temp_path.unlink(missing_ok=True)

    storage_module.records_path(target).unlink(missing_ok=True)
    registry = getattr(automation_module, "_automation_task_by_workspace", {})
    registry[str(workspace)] = task
    records = automation_module.load_records(target)
    return target, len(records)


def read_artifact_records(
    path: Path,
    task: Any | None,
    storage_module: Any,
) -> tuple[list[dict[str, Any]], bool]:
    suffix = Path(path).suffix.lower()
    if suffix == ".xlsx":
        rows, metadata, recognized = _read_xlsx(Path(path), task)
    elif suffix == ".csv":
        rows, metadata, recognized = _read_csv(Path(path))
    elif suffix == ".json":
        rows, metadata, recognized = _read_json(Path(path))
    elif suffix == ".md":
        rows, metadata, recognized = _read_markdown(Path(path))
    else:
        return [], False
    if not recognized:
        return [], False

    timestamp = _file_timestamp(Path(path))
    records: list[dict[str, Any]] = []
    for index, raw_values in enumerate(rows[:MAX_IMPORTED_ROWS], start=1):
        values = _normalize_values(raw_values, task, storage_module)
        if not any(not _is_empty(value) for value in values.values()):
            continue
        meta = metadata[index - 1] if index - 1 < len(metadata) else {}
        record_id = str(meta.get("record_id") or _stable_record_id(Path(path), index))
        source_ids = _string_list(meta.get("source_message_ids"))
        created_at = str(meta.get("created_at") or timestamp)
        updated_at = str(meta.get("updated_at") or timestamp)
        dedup_key = str(meta.get("dedup_key") or _dedup_key(task, values))
        records.append(
            {
                "record_id": record_id,
                "values": values,
                "source_message_ids": source_ids,
                "created_at": created_at,
                "updated_at": updated_at,
                "dedup_key": dedup_key,
            }
        )
    return records, True


def latest_task_artifact(task: Any, automation_module: Any, storage_module: Any) -> Path | None:
    del automation_module
    workspace = storage_module.task_workspace(task.task_id)
    suffix = f".{str(task.file_format).lower()}"
    candidates = [
        path
        for path in workspace.iterdir()
        if path.is_file()
        and path.suffix.lower() == suffix
        and not path.name.endswith(".records.json")
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def _sidecar_is_current(storage_module: Any, artifact: Path) -> bool:
    sidecar = storage_module.records_path(artifact)
    if not sidecar.is_file():
        return False
    try:
        return sidecar.stat().st_mtime >= artifact.stat().st_mtime
    except OSError:
        return False


def _read_xlsx(
    path: Path,
    task: Any | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], bool]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise AutomationFileImportError(
            "缺少 openpyxl，请先执行 pip install -r requirements.txt"
        ) from exc

    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001
        raise AutomationFileImportError(f"无法读取 Excel：{exc}") from exc

    visible = [sheet for sheet in workbook.worksheets if sheet.title != META_SHEET_NAME]
    if not visible:
        workbook.close()
        return [], [], False
    sheet_name = str(getattr(task, "sheet_name", "") or "")
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else visible[0]
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        workbook.close()
        return [], [], True
    headers = _headers(all_rows[0])
    if not headers:
        workbook.close()
        return [], [], True
    rows = [_row_mapping(headers, values) for values in all_rows[1 : MAX_IMPORTED_ROWS + 1]]

    metadata: list[dict[str, Any]] = []
    if META_SHEET_NAME in workbook.sheetnames:
        meta_rows = list(workbook[META_SHEET_NAME].iter_rows(values_only=True))
        if meta_rows:
            meta_headers = _headers(meta_rows[0])
            metadata = [_row_mapping(meta_headers, values) for values in meta_rows[1:]]
    workbook.close()
    return rows, metadata, True


def _read_csv(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], bool]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                if reader.fieldnames is None:
                    return [], [], True
                rows = [
                    {str(key or "").strip(): value for key, value in row.items() if str(key or "").strip()}
                    for _, row in zip(range(MAX_IMPORTED_ROWS), reader)
                ]
                return rows, [], True
        except UnicodeDecodeError as exc:
            last_error = exc
    raise AutomationFileImportError(f"无法读取 CSV 编码：{last_error}")


def _read_json(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], bool]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise AutomationFileImportError(f"无法读取 JSON：{exc}") from exc

    if isinstance(payload, list):
        items = payload
    elif isinstance(payload, dict):
        items = next(
            (
                payload[key]
                for key in ("records", "rows", "data", "items")
                if isinstance(payload.get(key), list)
            ),
            [payload],
        )
    else:
        return [], [], False

    rows = [
        {str(key): value for key, value in item.items()}
        for item in items[:MAX_IMPORTED_ROWS]
        if isinstance(item, dict)
    ]
    return rows, [], True


def _read_markdown(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], bool]:
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError as exc:
        raise AutomationFileImportError(f"无法读取 Markdown：{exc}") from exc

    for index in range(len(lines) - 1):
        header_cells = _split_markdown_row(lines[index])
        separator_cells = _split_markdown_row(lines[index + 1])
        if not header_cells or len(separator_cells) != len(header_cells):
            continue
        if not all(re.fullmatch(r":?-{3,}:?", cell.strip()) for cell in separator_cells):
            continue
        headers = [cell.strip() for cell in header_cells]
        rows: list[dict[str, Any]] = []
        for raw in lines[index + 2 :]:
            cells = _split_markdown_row(raw)
            if not cells:
                break
            cells += [""] * max(0, len(headers) - len(cells))
            rows.append(
                {
                    header: cells[position].replace("<br>", "\n")
                    for position, header in enumerate(headers)
                    if header
                }
            )
            if len(rows) >= MAX_IMPORTED_ROWS:
                break
        return rows, [], True
    return [], [], False


def _headers(row: Any) -> list[str]:
    values = list(row or [])
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = str(value or "").strip() or f"列{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    while headers and headers[-1].startswith("列") and _is_empty(values[len(headers) - 1]):
        headers.pop()
    return headers


def _row_mapping(headers: list[str], row: Any) -> dict[str, Any]:
    values = list(row or [])
    values += [None] * max(0, len(headers) - len(values))
    return {header: values[index] for index, header in enumerate(headers)}


def _normalize_values(raw: dict[str, Any], task: Any | None, storage_module: Any) -> dict[str, Any]:
    if task is None:
        return {str(key): value for key, value in raw.items()}
    result: dict[str, Any] = {}
    for column in task.columns:
        value = raw.get(column.name, column.default if column.default != "" else "")
        coerce = getattr(storage_module, "_coerce_value", None)
        result[column.name] = coerce(column, value) if callable(coerce) else value
    return result


def _dedup_key(task: Any | None, values: dict[str, Any]) -> str:
    fields = list(getattr(task, "dedup_fields", []) or []) if task is not None else []
    if not fields:
        return ""
    parts = [str(values.get(name, "")).strip().lower() for name in fields]
    if not any(parts):
        return ""
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def _stable_record_id(path: Path, row_index: int) -> str:
    key = f"{path.name}|{row_index}"
    return "row_import_" + hashlib.sha256(key.encode("utf-8")).hexdigest()[:16]


def _file_timestamp(path: Path) -> str:
    try:
        return datetime.fromtimestamp(path.stat().st_mtime).isoformat()
    except OSError:
        return datetime.now().isoformat()


def _string_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value if str(item)]
    text = str(value or "").strip()
    if not text:
        return []
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        return [part.strip() for part in re.split(r"[,，;；]", text) if part.strip()]
    return [str(item) for item in parsed if str(item)] if isinstance(parsed, list) else []


def _split_markdown_row(line: str) -> list[str]:
    text = str(line or "").strip()
    if not text.startswith("|") or not text.endswith("|"):
        return []
    body = text[1:-1]
    cells = re.split(r"(?<!\\)\|", body)
    return [cell.replace("\\|", "|").strip() for cell in cells]


def _is_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""
